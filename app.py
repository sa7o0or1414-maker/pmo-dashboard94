import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import timedelta
import altair as alt
import json
import io

try:
    import openai
except ImportError:
    openai = None

# ================= إعدادات الصفحة =================
st.set_page_config(
    page_title="لوحة المعلومات | PMO",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ================= Session State =================
defaults = {
    "role": "viewer",
    "page": "home",
    "show_overdue": False,
    "show_risk": False,
    "top_nav": "مشاريع الباب الثالث"
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

ADMIN_USER = "admin"
ADMIN_PASS = "1234"

DATA_DIR = Path("data")
ASSETS_DIR = Path("assets")
DATA_DIR.mkdir(exist_ok=True)
ASSETS_DIR.mkdir(exist_ok=True)

LOGO_WIDTH_FILE = Path("data/logo_width.txt")
LOGO_ALIGN_FILE = Path("data/logo_align.txt")
LOGO_EXCEL_FILE = Path("data/logo_excel.txt")
LOGO_EXCEL_WIDTH_FILE = Path("data/logo_excel_width.txt")
USERS_FILE = Path("data/users.json")
API_KEY_FILE = Path("data/api_key.txt")

# Load users
if USERS_FILE.exists():
    try:
        users = json.loads(USERS_FILE.read_text())
    except:
        users = {"admin": {"password": "1234", "role": "admin"}}
else:
    users = {"admin": {"password": "1234", "role": "admin"}}

# Load logo width
if LOGO_WIDTH_FILE.exists():
    try:
        logo_width = int(LOGO_WIDTH_FILE.read_text().strip())
    except:
        logo_width = 120
else:
    logo_width = 120

# Load logo alignment
if LOGO_ALIGN_FILE.exists():
    logo_alignment = LOGO_ALIGN_FILE.read_text().strip()
    if logo_alignment not in ["left", "center", "right"]:
        logo_alignment = "center"
else:
    logo_alignment = "center"

# Load show logo in excel
if LOGO_EXCEL_FILE.exists():
    show_logo_in_excel = LOGO_EXCEL_FILE.read_text().strip().lower() == "true"
else:
    show_logo_in_excel = True  # افتراضي True

# Load logo excel width
if LOGO_EXCEL_WIDTH_FILE.exists():
    try:
        logo_excel_width = int(LOGO_EXCEL_WIDTH_FILE.read_text().strip())
    except:
        logo_excel_width = 400
else:
    logo_excel_width = 400

# Load API key
if API_KEY_FILE.exists():
    api_key = API_KEY_FILE.read_text().strip()
else:
    api_key = ""

LOGO_PATH = ASSETS_DIR / "logo.png"
LOGO_EXCEL_PATH = ASSETS_DIR / "logo_excel.png"
TEMPLATE_PATH = ASSETS_DIR / "template.xlsx"

DATA_FILES = {
    "مشاريع الباب الثالث": "bab3.xlsx",
    "مشاريع الباب الرابع": "bab4.xlsx",
    "مشاريع بهجة": "bahja.xlsx",
    "تطبيق دليل PMD": "pmd.xlsx",
    "المشاريع المنجزة": "done.xlsx",
    "مشاريع المحفظة": "portfolio.xlsx",
    "الدراسات وقوائم التحقق": "studies.xlsx",
    "دورة المشتريات": "procurement.xlsx",
    "مواقع المشاريع": "sites.xlsx",
    "مشاريع الإسكان": "housing.xlsx",
    "الافتراضي": "data.xlsx"
}

# ================= CSS =================
st.markdown("""
<style>
html, body, [class*="css"] {
    direction: rtl;
    font-family: -apple-system, BlinkMacSystemFont, 'SF Pro Display', 'SF Pro Text', 'Helvetica Neue', sans-serif;
    line-height: 1.5;
    color: #1d1d1f;
}
@media (prefers-color-scheme: dark) {
    html, body, [class*="css"] {
        color: #f5f5f7;
        background-color: #000000;
    }
}

h1 {
    text-align: center;
    font-weight: 600;
    font-size: 2.5rem;
    margin-bottom: 2rem;
    color: #1d1d1f;
}
@media (prefers-color-scheme: dark) {
    h1 {
        color: #f5f5f7;
    }
}

h2, h3, h4, h5, h6 {
    text-align: center !important;
    color: #153e46 !important;
}

section[data-testid="stSidebar"] {
    background: #153e46;
    border-right: 1px solid #d2d2d7;
    box-shadow: 0 0 20px rgba(0,0,0,0.05);
    position: absolute;
    right: 0;
    top: 0;
    height: 100vh;
    transition: width 0.3s ease;
}
@media (prefers-color-scheme: dark) {
    section[data-testid="stSidebar"] {
        background: #153e46;
        border-right: 1px solid #424245;
        box-shadow: 0 0 20px rgba(0,0,0,0.2);
    }
}
section[data-testid="stSidebar"] * {
    color: #ffffff;
}
@media (prefers-color-scheme: dark) {
    section[data-testid="stSidebar"] * {
        color: #ffffff;
    }
}

section[data-testid="stSidebar"] button {
    width: 120px !important;
    height: 50px !important;
    border-radius: 8px !important;
    margin: 10px auto !important;
    display: block !important;
    background: rgba(255,255,255,0.08) !important;
    border: 2px solid rgba(255,255,255,0.3) !important;
    color: #ffffff !important;
    font-size: 14px !important;
    text-align: center !important;
    padding: 0 !important;
    line-height: 46px !important; /* adjusted for border */
    font-weight: 700 !important;
    transition: all 0.3s ease !important;
    box-shadow: 0 2px 4px rgba(0,0,0,0.2) !important;
}
section[data-testid="stSidebar"] button:hover {
    background: rgba(255,255,255,0.2) !important;
    border-color: rgba(255,255,255,0.6) !important;
    box-shadow: 0 4px 12px rgba(0,0,0,0.4) !important;
    transform: translateY(-2px) !important;
}

section[data-testid="stSidebar"]:not([data-expanded="true"]) {
    width: 0 !important;
    overflow: hidden !important;
    padding: 0 !important;
    margin: 0 !important;
    border: none !important;
}

.card {
    background: #ffffff;
    padding: 24px;
    border-radius: 16px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.05);
    text-align: center;
    min-height: 140px;
    border: 1px solid #d2d2d7;
    transition: all 0.3s ease;
}
@media (prefers-color-scheme: dark) {
    .card {
        background: #1d1d1f;
        border: 1px solid #424245;
        box-shadow: 0 2px 10px rgba(0,0,0,0.2);
    }
}
.card:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 20px rgba(0,0,0,0.1);
}
@media (prefers-color-scheme: dark) {
    .card:hover {
        box-shadow: 0 4px 20px rgba(0,0,0,0.3);
    }
}
.card h2 {
    font-size: 24px;
    margin-bottom: 8px;
    font-weight: 600;
    color: #1d1d1f;
}
@media (prefers-color-scheme: dark) {
    .card h2 {
        color: #f5f5f7;
    }
}
.card.blue { border-top: 4px solid #007aff; }
.card.green { border-top: 4px solid #34c759; }
.card.orange { border-top: 4px solid #ff9500; }
.card.gray { border-top: 4px solid #8e8e93; }

.chart-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
    gap: 20px;
}
.chart-item {
    background: #ffffff;
    padding: 16px;
    border-radius: 8px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    transition: all 0.3s ease;
}
@media (prefers-color-scheme: dark) {
    .chart-item {
        background: #1d1d1f;
        box-shadow: 0 2px 8px rgba(0,0,0,0.2);
    }
}
.chart-item:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 16px rgba(0,0,0,0.15);
}
@media (prefers-color-scheme: dark) {
    .chart-item:hover {
        box-shadow: 0 4px 16px rgba(0,0,0,0.3);
    }
}

.topbar-btn button {
    background: transparent !important;
    border: 1px solid #d2d2d7 !important;
    border-radius: 12px !important;
    padding: 12px 20px !important;
    font-weight: 500;
    transition: all 0.2s ease;
    color: #007aff;
}
@media (prefers-color-scheme: dark) {
    .topbar-btn button {
        border: 1px solid #424245 !important;
        color: #0a84ff;
    }
}
.topbar-btn button:hover {
    background: #f5f5f7 !important;
    border-color: #007aff !important;
}
@media (prefers-color-scheme: dark) {
    .topbar-btn button:hover {
        background: #2c2c2e !important;
        border-color: #0a84ff !important;
    }
}

/* Mobile-friendly styles */
@media (max-width: 768px) {
    .card {
        padding: 16px;
        min-height: 120px;
        margin-bottom: 16px;
    }
    .card h2 { font-size: 20px; }
    h1 { font-size: 2rem; margin-bottom: 1.5rem; }
    .topbar-btn button {
        padding: 10px 16px !important;
        font-size: 14px;
    }
    .stHorizontalBlock {
        flex-direction: row !important;
        flex-wrap: wrap !important;
        gap: 12px !important;
    }
    .stHorizontalBlock > div {
        flex: 1 1 45% !important;
        min-width: 140px !important;
        margin-bottom: 12px !important;
    }
    .stDataFrame {
        overflow-x: auto !important;
        font-size: 12px !important;
    }
    .stSelectbox, .stTextInput {
        font-size: 14px !important;
    }
    .stButton button {
        font-size: 14px !important;
        padding: 10px 16px !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ================= أدوات =================
def load_data():
    file = DATA_FILES.get(st.session_state.top_nav, "data.xlsx")
    path = DATA_DIR / file
    if not path.exists():
        return None

    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    df.rename(columns={
        "إسم المشـــروع": "اسم المشروع",
        "قيمة المستخلصات المعتمده": "قيمة المستخلصات",
        "تاريخ الانتهاء من المشروع": "تاريخ الانتهاء",
    }, inplace=True)

    for c in ["قيمة العقد","قيمة المستخلصات","نسبة الإنجاز","نسبة الانجاز"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    if "السنة" in df.columns:
        df["السنة"] = pd.to_numeric(df["السنة"], errors="coerce").astype("Int64")

    if "تاريخ الانتهاء" in df.columns:
        df["تاريخ الانتهاء"] = pd.to_datetime(df["تاريخ الانتهاء"], errors="coerce")

    return df


def status_color(s):
    s = str(s)
    if "متأخر" in s or "متعثر" in s: return "#e63946"
    if "مكتمل" in s or "منجز" in s: return "#00a389"
    if "جاري" in s or "قيد" in s: return "#2c7be5"
    if "منتظ" in s: return "#34c759"
    return "#f4a261"


def build_status_df(df):
    s = df["حالة المشروع"].fillna("غير محدد").astype(str)
    out = s.value_counts().rename_axis("الحالة").reset_index(name="عدد")
    out["لون"] = out["الحالة"].apply(status_color)
    return out


def create_excel_from_template(filtered_df, template_path, logo_path, show_logo, logo_width):
    # Fill NaN values to avoid Excel conversion errors
    filtered_df = filtered_df.astype(object).fillna('')

    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import PatternFill, Font

    n_cols = len(filtered_df.columns)
    last_col_letter = chr(64 + n_cols)

    if template_path.exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active

    # دمج الأسطر الأولى على عرض الجدول
    ws.merge_cells(f'A1:{last_col_letter}4')

    # إضافة اللوجو إذا كان مطلوباً
    if show_logo:
        if logo_path.exists():
            img_path = logo_path
        elif LOGO_PATH.exists():
            img_path = LOGO_PATH
        else:
            img_path = None
        if img_path:
            img = Image(img_path)
            img.width = logo_width
            img.height = logo_width // 4
            ws.add_image(img, 'B2')  # وضع في الوسط

    # إضافة عناوين الأعمدة في صف 5 بلون اللوجو والنص أبيض
    header_row = 5
    logo_fill = PatternFill(start_color="153E46", end_color="153E46", fill_type="solid")
    white_font = Font(color="FFFFFF")
    for c, header in enumerate(filtered_df.columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=header)
        cell.fill = logo_fill
        cell.font = white_font

    # إضافة البيانات بدءاً من صف 6
    start_row = 6
    for r, row in enumerate(filtered_df.itertuples(index=False), start=start_row):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def simple_chatbot_response(prompt, context):
    prompt_lower = prompt.lower()
    df = context['filtered']

    # أسئلة عن البيانات والمشاريع
    if "عدد" in prompt_lower and "مشروع" in prompt_lower:
        project_count = len(df)
        status_counts = df["حالة المشروع"].value_counts() if "حالة المشروع" in df.columns else {}
        response = f"📊 **عدد المشاريع الحالي: {project_count}**\n\n"
        if not status_counts.empty:
            response += "**توزيع حسب الحالة:**\n"
            for status, count in status_counts.items():
                response += f"• {status}: {count}\n"
        return response

    elif "قيمة" in prompt_lower and "عقد" in prompt_lower:
        total_contract = context['total_contract']
        if "قيمة العقد" in df.columns:
            top_projects = df.nlargest(5, "قيمة العقد")[["اسم المشروع", "قيمة العقد"]] if "اسم المشروع" in df.columns else df.nlargest(5, "قيمة العقد")[["قيمة العقد"]]
            response = f"💰 **قيمة العقود الإجمالية: {total_contract:,.0f} ريال**\n\n"
            response += "**أكبر 5 مشاريع قيمة:**\n"
            for _, row in top_projects.iterrows():
                if "اسم المشروع" in row:
                    response += f"• {row['اسم المشروع']}: {row['قيمة العقد']:,.0f} ريال\n"
                else:
                    response += f"• {row['قيمة العقد']:,.0f} ريال\n"
        else:
            response = f"💰 **قيمة العقود الإجمالية: {total_contract:,.0f} ريال**"
        return response

    elif "نسبة" in prompt_lower and "إنجاز" in prompt_lower:
        progress_ratio = context['progress_ratio']
        if "نسبة الإنجاز" in df.columns and "قيمة العقد" in df.columns:
            completed_projects = df[df["نسبة الإنجاز"] >= 100]
            ongoing_projects = df[(df["نسبة الإنجاز"] > 0) & (df["نسبة الإنجاز"] < 100)]
            not_started = df[df["نسبة الإنجاز"] == 0]
            response = f"📈 **نسبة الإنجاز العامة: {progress_ratio:.1f}%**\n\n"
            response += f"• مشاريع مكتملة (100%): {len(completed_projects)}\n"
            response += f"• مشاريع قيد التنفيذ: {len(ongoing_projects)}\n"
            response += f"• مشاريع لم تبدأ: {len(not_started)}\n"
        else:
            response = f"📈 **نسبة الإنجاز العامة: {progress_ratio:.1f}%**"
        return response

    elif "متأخر" in prompt_lower or "متعثر" in prompt_lower:
        overdue_count = len(df[df["حالة المشروع"].astype(str).str.contains("متأخر|متعثر", na=False)])
        if overdue_count > 0 and "اسم المشروع" in df.columns:
            overdue_projects = df[df["حالة المشروع"].astype(str).str.contains("متأخر|متعثر", na=False)]["اسم المشروع"].head(10)
            response = f"⚠️ **عدد المشاريع المتأخرة/المتعثرة: {overdue_count}**\n\n"
            response += "**أسماء المشاريع المتأخرة (أول 10):**\n"
            for name in overdue_projects:
                response += f"• {name}\n"
        else:
            response = f"⚠️ **عدد المشاريع المتأخرة/المتعثرة: {overdue_count}**"
        return response

    elif "منجز" in prompt_lower or "مكتمل" in prompt_lower:
        completed_count = len(df[df["حالة المشروع"].astype(str).str.contains("منجز|مكتمل|منتهي", na=False)])
        if completed_count > 0 and "اسم المشروع" in df.columns:
            completed_projects = df[df["حالة المشروع"].astype(str).str.contains("منجز|مكتمل|منتهي", na=False)]["اسم المشروع"].head(10)
            response = f"✅ **عدد المشاريع المنجزة: {completed_count}**\n\n"
            response += "**أسماء المشاريع المنجزة (أول 10):**\n"
            for name in completed_projects:
                response += f"• {name}\n"
        else:
            response = f"✅ **عدد المشاريع المنجزة: {completed_count}**"
        return response

    elif "جاري" in prompt_lower or "قيد التنفيذ" in prompt_lower:
        ongoing_count = len(df[df["حالة المشروع"].astype(str).str.contains("جاري|قيد التنفيذ|نشط", na=False)])
        if ongoing_count > 0 and "اسم المشروع" in df.columns:
            ongoing_projects = df[df["حالة المشروع"].astype(str).str.contains("جاري|قيد التنفيذ|نشط", na=False)]["اسم المشروع"].head(10)
            response = f"🔄 **عدد المشاريع قيد التنفيذ: {ongoing_count}**\n\n"
            response += "**أسماء المشاريع قيد التنفيذ (أول 10):**\n"
            for name in ongoing_projects:
                response += f"• {name}\n"
        else:
            response = f"🔄 **عدد المشاريع قيد التنفيذ: {ongoing_count}**"
        return response

    elif "بلدية" in prompt_lower and "عدد" in prompt_lower:
        if "البلدية" in df.columns:
            municipal_counts = df["البلدية"].value_counts().head(10)
            response = "🏛️ **عدد المشاريع حسب البلدية:**\n\n"
            for municipal, count in municipal_counts.items():
                response += f"• {municipal}: {count} مشروع\n"
            return response
        else:
            return "🏛️ لا توجد بيانات البلديات المتاحة"

    elif "أكبر" in prompt_lower and "قيمة" in prompt_lower:
        if "قيمة العقد" in df.columns:
            max_contract = df["قيمة العقد"].max()
            project_name = df.loc[df["قيمة العقد"].idxmax(), "اسم المشروع"] if "اسم المشروع" in df.columns else "غير محدد"
            response = f"💎 **المشروع الأكبر قيمة:**\n"
            response += f"• اسم المشروع: {project_name}\n"
            response += f"• القيمة: {max_contract:,.0f} ريال\n"
            if "البلدية" in df.columns:
                municipal = df.loc[df["قيمة العقد"].idxmax(), "البلدية"]
                response += f"• البلدية: {municipal}\n"
            if "حالة المشروع" in df.columns:
                status = df.loc[df["قيمة العقد"].idxmax(), "حالة المشروع"]
                response += f"• الحالة: {status}\n"
        else:
            response = "💎 لا توجد بيانات قيم العقود"
        return response

    elif "قائمة" in prompt_lower and "مشاريع" in prompt_lower:
        if "اسم المشروع" in df.columns:
            projects_list = df["اسم المشروع"].head(20).tolist()
            response = f"📋 **قائمة المشاريع (أول 20):**\n\n"
            for i, name in enumerate(projects_list, 1):
                response += f"{i}. {name}\n"
            if len(df) > 20:
                response += f"\n... و {len(df) - 20} مشروع آخر"
        else:
            response = "📋 لا توجد بيانات أسماء المشاريع"
        return response

    elif "تحليل" in prompt_lower and "كامل" in prompt_lower:
        response = "📊 **تحليل شامل للبيانات:**\n\n"
        response += f"• إجمالي المشاريع: {len(df)}\n"
        response += f"• إجمالي قيمة العقود: {context['total_contract']:,.0f} ريال\n"
        response += f"• متوسط نسبة الإنجاز: {context['progress_ratio']:.1f}%\n"
        response += f"• عدد المشاريع المتأخرة: {len(df[df['حالة المشروع'].astype(str).str.contains('متأخر|متعثر', na=False)])}\n"
        response += f"• عدد المشاريع المنجزة: {len(df[df['حالة المشروع'].astype(str).str.contains('منجز|مكتمل|منتهي', na=False)])}\n"
        response += f"• عدد المشاريع قيد التنفيذ: {len(df[df['حالة المشروع'].astype(str).str.contains('جاري|قيد التنفيذ|نشط', na=False)])}\n"
        if "البلدية" in df.columns:
            top_municipal = df["البلدية"].value_counts().head(3)
            response += "\n**أكثر البلديات نشاطاً:**\n"
            for municipal, count in top_municipal.items():
                response += f"• {municipal}: {count} مشروع\n"
        return response

    # أسئلة عن الموقع والاستخدام
    elif "كيف" in prompt_lower and ("استخدم" in prompt_lower or "استخدام" in prompt_lower):
        return """لاستخدام الموقع:
• اختر نوع المشاريع من الأزرار العلوية
• استخدم الفلاتر لتصفية البيانات حسب البلدية، الجهة، الحالة، إلخ
• اضغط على 'إعادة تعيين الفلاتر' لإزالة جميع الفلاتر
• استخدم الدردشة للسؤال عن البيانات
• اضغط على 'تحميل البيانات كExcel' لتصدير البيانات المفلترة"""

    elif "ما هي" in prompt_lower and ("بيانات" in prompt_lower or "معلومات" in prompt_lower):
        return """البيانات المتاحة تشمل:
• مشاريع الباب الثالث والرابع
• مشاريع بهجة
• تطبيق دليل PMD
• المشاريع المنجزة
• مشاريع المحفظة
• الدراسات وقوائم التحقق
• دورة المشتريات
• مواقع المشاريع
• مشاريع الإسكان"""

    elif "كيف" in prompt_lower and ("تصفية" in prompt_lower or "فلتر" in prompt_lower):
        return """لتصفية البيانات:
• اختر نوع المشاريع من الأزرار العلوية
• استخدم القوائم المنسدلة لتحديد البلدية، الجهة، الحالة، إلخ
• الفلاتر تتغير ديناميكياً حسب اختياراتك السابقة
• اضغط 'إعادة تعيين الفلاتر' لإزالة جميع الفلاتر"""

    elif "ما هي" in prompt_lower and ("مؤشرات" in prompt_lower or "kpi" in prompt_lower):
        return """المؤشرات الرئيسية المعروضة:
• عدد المشاريع
• قيمة العقود الإجمالية
• قيمة المستخلصات
• المبلغ المتبقي من المستخلصات
• نسبة الصرف
• نسبة الإنجاز العامة"""

    elif "كيف" in prompt_lower and ("تحميل" in prompt_lower or "تصدير" in prompt_lower):
        return """لتحميل البيانات:
• قم بتصفية البيانات حسب الحاجة
• اضغط على 'تحميل البيانات كExcel'
• سيتم تحميل ملف Excel يحتوي على البيانات المفلترة مع الشعار"""

    elif "ما هي" in prompt_lower and ("تنبيهات" in prompt_lower or "مشاريع متأخرة" in prompt_lower):
        return """التنبيهات تشمل:
• المشاريع المتأخرة أو المتعثرة
• المشاريع المتوقع تأخرها (قريبة من التاريخ المحدد ومنخفضة الإنجاز)
• يمكنك عرض التفاصيل والتحميل كملفات Excel منفصلة"""

    elif "كيف" in prompt_lower and ("تسجيل" in prompt_lower or "دخول" in prompt_lower):
        return """للتسجيل الدخول:
• اضغط على 'تسجيل الدخول' في الشريط الجانبي
• أدخل اسم المستخدم وكلمة المرور
• المستخدم الافتراضي: admin / 1234
• المدراء يمكنهم الوصول للإعدادات ورفع البيانات"""

    elif "ما هي" in prompt_lower and ("إعدادات" in prompt_lower or "اعدادات" in prompt_lower):
        return """الإعدادات المتاحة للمدراء:
• إدارة المستخدمين (إضافة/حذف)
• رفع وتخصيص الشعار
• إعدادات ملفات Excel المُحمّلة
• معلومات الدردشة"""

    elif "كيف" in prompt_lower and ("رفع" in prompt_lower and "بيانات" in prompt_lower):
        return """لرفع البيانات:
• سجل الدخول كمدير
• اضغط على 'رفع البيانات' في الشريط الجانبي
• اختر نوع المشاريع وارفع ملف Excel الجديد
• سيتم استبدال البيانات القديمة بالجديدة"""

    elif "ما هي" in prompt_lower and ("أقسام" in prompt_lower or "اقسام" in prompt_lower):
        return """أقسام الموقع:
• الصفحة الرئيسية: عرض البيانات والتحليلات
• الدردشة: طرح الأسئلة عن البيانات
• تسجيل الدخول: للمدراء
• الإعدادات: تخصيص الموقع (للمدراء)
• رفع البيانات: تحديث ملفات البيانات (للمدراء)"""

    elif "كيف" in prompt_lower and ("دردشة" in prompt_lower or "سؤال" in prompt_lower):
        return """لاستخدام الدردشة:
• اضغط على '🤖 اسألني' في الشريط الجانبي
• اكتب سؤالك باللغة العربية
• يمكنك السؤال عن عدد المشاريع، القيم، النسب، إلخ
• الدردشة تعمل بدون الحاجة لمفتاح API"""

    elif "من" in prompt_lower and ("طور" in prompt_lower or "صنع" in prompt_lower):
        return "تم تطوير هذا الموقع بواسطة فريق PMO لإدارة وتحليل مشاريع البلدية بطريقة احترافية وسهلة الاستخدام."

    elif "ما هي" in prompt_lower and ("ميزات" in prompt_lower or "خصائص" in prompt_lower):
        return """ميزات الموقع:
• واجهة عربية مع دعم RTL
• تحليلات بصرية متقدمة
• فلاتر ديناميكية
• تصدير البيانات كملفات Excel
• نظام دردشة ذكي
• إدارة المستخدمين
• تنبيهات المشاريع
• تصميم متجاوب للهواتف"""

    else:
        return """أنا مساعد ذكي لموقع لوحة معلومات PMO. يمكنني المساعدة في:

📊 **الأسئلة عن البيانات:**
• عدد المشاريع، قيمة العقود، نسبة الإنجاز
• المشاريع المتأخرة، المنجزة، قيد التنفيذ
• توزيع المشاريع حسب البلدية
• أكبر المشاريع قيمة
• قائمة المشاريع
• تحليل شامل

🛠️ **الأسئلة عن الموقع:**
• كيفية الاستخدام والتصفية
• المؤشرات والتنبيهات
• التسجيل الدخول والإعدادات
• رفع البيانات والتصدير

💡 **نصائح:**
• جرب أسئلة مثل: "عدد المشاريع"، "قائمة المشاريع"، "تحليل كامل"، "كيف أستخدم الموقع"
• يمكنك السؤال باللغة العربية الطبيعية

اسأل عن أي شيء يخص الموقع أو البيانات!"""

# ================= Sidebar =================
with st.sidebar:
    if LOGO_PATH.exists():
        if logo_alignment == "center":
            col1, col2, col3 = st.columns([1, 1, 1])
            with col2:
                st.image(LOGO_PATH, width=logo_width)
        elif logo_alignment == "right":
            col1, col2 = st.columns([1, 1])
            with col2:
                st.image(LOGO_PATH, width=logo_width)
        else:  # left
            st.image(LOGO_PATH, width=logo_width)

    if st.button("الصفحة الرئيسية"):
        st.session_state.page = "home"

    if st.button("🤖 اسألني"):
        st.session_state.page = "chat"

    if st.session_state.role == "viewer":
        if st.button("تسجيل الدخول"):
            st.session_state.page = "login"

    if st.session_state.role == "admin":
        if st.button("⚙️ الإعدادات"):
            st.session_state.page = "settings"
        if st.button("رفع البيانات"):
            st.session_state.page = "upload"
        if st.button("تسجيل خروج"):
            st.session_state.role = "viewer"
            st.session_state.page = "home"
            st.rerun()

# ================= Login =================
if st.session_state.page == "login":
    st.title("تسجيل الدخول")
    u = st.text_input("اسم المستخدم")
    p = st.text_input("كلمة المرور", type="password")
    if st.button("دخول"):
        if u in users and users[u]["password"] == p:
            st.session_state.role = users[u]["role"]
            st.session_state.page = "home"
            st.rerun()
        else:
            st.error("بيانات غير صحيحة")
    st.stop()

# ================= Upload =================
if st.session_state.page == "upload":
    st.title("رفع الملفات حسب نوع المشاريع")
    for name, file in DATA_FILES.items():
        if name == "الافتراضي":
            continue
        with st.expander(name):
            up = st.file_uploader(name, type=["xlsx"], key=file)
            if up:
                (DATA_DIR / file).write_bytes(up.getbuffer())
                st.success("تم رفع الملف")
    st.stop()

# ================= Settings =================
if st.session_state.page == "settings":
    st.title("الإعدادات")

    # User Management
    st.subheader("إدارة المستخدمين")
    st.write("المستخدمون الحاليون:")
    for user, data in users.items():
        st.write(f"- {user}: {data['role']}")

    with st.expander("إضافة مستخدم جديد"):
        new_user = st.text_input("اسم المستخدم الجديد")
        new_pass = st.text_input("كلمة المرور", type="password")
        new_role = st.selectbox("الدور", ["viewer", "admin"])
        if st.button("إضافة المستخدم"):
            if new_user and new_pass:
                users[new_user] = {"password": new_pass, "role": new_role}
                USERS_FILE.write_text(json.dumps(users, ensure_ascii=False, indent=2))
                st.success("تم إضافة المستخدم")
                st.rerun()
            else:
                st.error("يرجى ملء جميع الحقول")

    # Logo Settings
    st.subheader("إعدادات الشعار")
    logo_upload = st.file_uploader("رفع شعار جديد", type=["png", "jpg", "jpeg"])
    if logo_upload:
        LOGO_PATH.write_bytes(logo_upload.getbuffer())
        st.success("تم رفع الشعار")

    current_width = st.slider("عرض الشعار", 50, 200, logo_width)
    if current_width != logo_width:
        LOGO_WIDTH_FILE.write_text(str(current_width))
        st.success("تم حفظ العرض")
        st.rerun()

    current_align = st.selectbox("محاذاة الشعار", ["left", "center", "right"], index=["left", "center", "right"].index(logo_alignment))
    if current_align != logo_alignment:
        LOGO_ALIGN_FILE.write_text(current_align)
        st.success("تم حفظ المحاذاة")
        st.rerun()

    st.subheader("إعدادات ملفات Excel المُحمّلة")
    logo_excel_upload = st.file_uploader("رفع لوجو لملفات Excel", type=["png", "jpg", "jpeg"], key="logo_excel")
    if logo_excel_upload:
        LOGO_EXCEL_PATH.write_bytes(logo_excel_upload.getbuffer())
        st.success("تم رفع لوجو Excel")

    current_excel_width = st.slider("عرض اللوجو في Excel", 200, 800, logo_excel_width)
    if current_excel_width != logo_excel_width:
        LOGO_EXCEL_WIDTH_FILE.write_text(str(current_excel_width))
        st.success("تم حفظ عرض اللوجو في Excel")
        st.rerun()

    show_logo_excel = st.checkbox("إظهار اللوجو في ملفات Excel المُحمّلة", value=show_logo_in_excel)
    if show_logo_excel != show_logo_in_excel:
        LOGO_EXCEL_FILE.write_text(str(show_logo_excel))
        st.success("تم حفظ إعداد اللوجو في Excel")
        st.rerun()

    st.subheader("إعدادات الدردشة")
    st.write("الدردشة متاحة للجميع بدون الحاجة إلى مفتاح API.")

    st.stop()

# ================= Chat =================
if st.session_state.page == "chat":
    st.title("🤖 اسألني")

    # Load data for context
    df_chat = load_data()
    if df_chat is None:
        st.warning("لا توجد بيانات متاحة.")
        st.stop()

    # Simple filtering for context (can be enhanced)
    filtered_chat = df_chat.copy()
    total_contract_chat = filtered_chat["قيمة العقد"].sum() if "قيمة العقد" in filtered_chat.columns else 0
    progress_ratio_chat = 0
    if "نسبة الإنجاز" in filtered_chat.columns and "قيمة العقد" in filtered_chat.columns:
        w = filtered_chat.dropna(subset=["قيمة العقد","نسبة الإنجاز"])
        if not w.empty:
            progress_ratio_chat = (w["قيمة العقد"] * w["نسبة الإنجاز"]).sum() / w["قيمة العقد"].sum()

    context = {
        'filtered': filtered_chat,
        'total_contract': total_contract_chat,
        'progress_ratio': progress_ratio_chat
    }

    if "messages" not in st.session_state:
        st.session_state.messages = []

    # عرض الرسائل السابقة
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    # إدخال السؤال
    if prompt := st.chat_input("اسأل عن التحليل أو اللوحة..."):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        # استخدام الchatbot البسيط
        answer = simple_chatbot_response(prompt, context)

        st.session_state.messages.append({"role": "assistant", "content": answer})
        with st.chat_message("assistant"):
            st.markdown(answer)

    st.stop()

# ================= Home =================
st.title("لوحة المعلومات")

# ===== Top Buttons =====
items = list(DATA_FILES.keys())
items.remove("الافتراضي")

r1 = st.columns(5)
for name, col in zip(items[:5], r1):
    with col:
        st.markdown("<div class='topbar-btn'>", unsafe_allow_html=True)
        if st.button(name):
            st.session_state.top_nav = name
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

r2 = st.columns(5)
for name, col in zip(items[5:], r2):
    with col:
        st.markdown("<div class='topbar-btn'>", unsafe_allow_html=True)
        if st.button(name):
            st.session_state.top_nav = name
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

st.caption(f"📊 التحليل الحالي: {st.session_state.top_nav}")

df = load_data()
if df is None:
    st.warning("لا يوجد ملف لهذا القسم")
    st.stop()

# ================= تحليل خاص لمشاريع بهجة =================
if st.session_state.top_nav == "مشاريع بهجة":
    st.subheader("تحليل مشاريع بهجة")

    # Initialize filter states as lists
    if "bahja_mun" not in st.session_state: st.session_state.bahja_mun = []
    if "bahja_project" not in st.session_state: st.session_state.bahja_project = []
    if "bahja_ptype" not in st.session_state: st.session_state.bahja_ptype = []

    if st.button("إعادة تعيين الفلاتر"):
        st.session_state.bahja_mun = []
        st.session_state.bahja_project = []
        st.session_state.bahja_ptype = []
        st.rerun()

    # Define filter columns and keys
    filter_cols = {
        "البلدية": "bahja_mun",
        "اسم المشروع": "bahja_project",
        "نوع المشروع": "bahja_ptype"
    }

    def get_filtered_excluding_bahja(df, filter_cols, exclude_col=None):
        temp = df.copy()
        for col, key in filter_cols.items():
            if col != exclude_col:
                sel = st.session_state[key]
                if isinstance(sel, str):
                    if sel != "الكل" and col in temp.columns:
                        temp = temp[temp[col] == sel]
                elif sel and col in temp.columns:
                    temp = temp[temp[col].isin(sel)]
        return temp

    f1,f2,f3 = st.columns(3)

    # Build options for each filter from data filtered by others
    mun_filtered = get_filtered_excluding_bahja(df, filter_cols, "البلدية")
    mun_options = sorted(mun_filtered["البلدية"].dropna().unique()) if "البلدية" in mun_filtered.columns else []
    mun_default = st.session_state.bahja_mun if isinstance(st.session_state.bahja_mun, list) else ([st.session_state.bahja_mun] if st.session_state.bahja_mun != "الكل" else [])
    mun = f1.multiselect("البلدية", mun_options, default=mun_default, key="bahja_mun")

    project_filtered = get_filtered_excluding_bahja(df, filter_cols, "اسم المشروع")
    project_options = sorted(project_filtered["اسم المشروع"].dropna().unique()) if "اسم المشروع" in project_filtered.columns else []
    project_default = st.session_state.bahja_project if isinstance(st.session_state.bahja_project, list) else ([st.session_state.bahja_project] if st.session_state.bahja_project != "الكل" else [])
    project = f2.multiselect("اسم المشروع", project_options, default=project_default, key="bahja_project")

    ptype_filtered = get_filtered_excluding_bahja(df, filter_cols, "نوع المشروع")
    ptype_options = sorted(ptype_filtered["نوع المشروع"].dropna().unique()) if "نوع المشروع" in ptype_filtered.columns else []
    ptype_default = st.session_state.bahja_ptype if isinstance(st.session_state.bahja_ptype, list) else ([st.session_state.bahja_ptype] if st.session_state.bahja_ptype != "الكل" else [])
    ptype = f3.multiselect("نوع المشروع", ptype_options, default=ptype_default, key="bahja_ptype")

    # Apply all filters
    filtered = get_filtered_excluding_bahja(df, filter_cols)

    total_cost = filtered["التكلفة"].sum()
    progress_col = "نسبة الإنجاز" if "نسبة الإنجاز" in filtered.columns else "نسبة الانجاز"
    avg_progress = pd.to_numeric(filtered[progress_col], errors="coerce").mean()

    c1,c2,c3 = st.columns(3)
    c1.markdown(f"<div class='card blue'><h2>{len(filtered)}</h2>عدد المشاريع</div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='card green'><h2>{total_cost:,.0f}</h2>إجمالي التكلفة</div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='card orange'><h2>{avg_progress:.1f}%</h2>نسبة الإنجاز</div>", unsafe_allow_html=True)

    st.markdown('<div class="chart-grid">', unsafe_allow_html=True)

    # Chart 1
    st.markdown('<div class="chart-item">', unsafe_allow_html=True)
    st.subheader("حالة المشروع")
    st.bar_chart(filtered["حالة المشروع"].value_counts())
    st.markdown('</div>', unsafe_allow_html=True)

    # Chart 2
    st.markdown('<div class="chart-item">', unsafe_allow_html=True)
    st.subheader("المستهدف")
    st.bar_chart(filtered["المستهدف"].value_counts())
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    st.subheader("تفاصيل مشاريع بهجة")
    st.dataframe(filtered, use_container_width=True)

    # زر تحميل البيانات المفلترة كملف Excel باستخدام القالب
    excel_data = create_excel_from_template(filtered, TEMPLATE_PATH, LOGO_EXCEL_PATH, show_logo_in_excel, logo_excel_width)
    st.download_button(
        label="تحميل البيانات كExcel",
        data=excel_data,
        file_name=f"{st.session_state.top_nav.replace(' ', '_')}_filtered.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.stop()


# ================= تحليل خاص لتطبيق دليل PMD =================
if st.session_state.top_nav == "تطبيق دليل PMD":
    st.subheader("تحليل تطبيق دليل PMD")

    # Initialize filter states as list
    if "pmd_mun" not in st.session_state: st.session_state.pmd_mun = []

    if st.button("إعادة تعيين الفلاتر"):
        st.session_state.pmd_mun = []
        st.rerun()

    f1 = st.columns(1)[0]

    # Build filter options
    mun_options = sorted(df["البلدية"].dropna().unique()) if "البلدية" in df.columns else []
    mun_default = st.session_state.pmd_mun if isinstance(st.session_state.pmd_mun, list) else ([st.session_state.pmd_mun] if st.session_state.pmd_mun != "الكل" else [])
    mun = f1.multiselect("البلدية", mun_options, default=mun_default, key="pmd_mun")

    # Apply filter
    filtered = df[df["البلدية"].isin(mun)] if mun else df

    # Calculate metrics
    num_projects = len(filtered)
    avg_application = pd.to_numeric(filtered.get("نسبة التطبيق", pd.Series()), errors="coerce").mean() if "نسبة التطبيق" in filtered.columns else 0
    avg_maturity = pd.to_numeric(filtered.get("نسبة النضج", pd.Series()), errors="coerce").mean() if "نسبة النضج" in filtered.columns else 0

    c1,c2,c3 = st.columns(3)
    c1.markdown(f"<div class='card blue'><h2>{num_projects}</h2>عدد المشاريع</div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='card green'><h2>{avg_application:.1f}%</h2>نسبة التطبيق</div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='card orange'><h2>{avg_maturity:.1f}%</h2>نسبة النضج</div>", unsafe_allow_html=True)

    st.subheader("تفاصيل تطبيق دليل PMD")
    # Show table with specific columns, ensuring اسم المشروع and المقاول are first
    if "اسم المشروع" in filtered.columns and "المقاول" in filtered.columns:
        display_cols = ["اسم المشروع", "المقاول"] + [col for col in filtered.columns if col not in ["اسم المشروع", "المقاول"]]
    else:
        display_cols = filtered.columns
    st.dataframe(filtered[display_cols], use_container_width=True)

    # زر تحميل البيانات المفلترة كملف Excel باستخدام القالب
    excel_data = create_excel_from_template(filtered, TEMPLATE_PATH, LOGO_EXCEL_PATH, show_logo_in_excel, logo_excel_width)
    st.download_button(
        label="تحميل البيانات كExcel",
        data=excel_data,
        file_name=f"{st.session_state.top_nav.replace(' ', '_')}_filtered.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.stop()


# ================= تحليل خاص للمشاريع المنجزة =================
if st.session_state.top_nav == "المشاريع المنجزة":
    st.subheader("تحليل المشاريع المنجزة")

    # Initialize filter states as lists for multiselect
    if "done_cat" not in st.session_state: st.session_state.done_cat = []
    if "done_mun" not in st.session_state: st.session_state.done_mun = []
    if "done_budget" not in st.session_state: st.session_state.done_budget = []
    if "done_year" not in st.session_state: st.session_state.done_year = []
    if "done_project" not in st.session_state: st.session_state.done_project = []

    if st.button("إعادة تعيين الفلاتر"):
        st.session_state.done_cat = []
        st.session_state.done_mun = []
        st.session_state.done_budget = []
        st.session_state.done_year = []
        st.session_state.done_project = []
        st.rerun()

    # Define filter columns and keys
    filter_cols = {
        "التصنيف": "done_cat",
        "البلدية": "done_mun",
        "ميزانية المشروع": "done_budget",
        "السنة": "done_year",
        "اسم المشروع": "done_project"
    }

    def get_filtered_excluding(df, filter_cols, exclude_col=None):
        temp = df.copy()
        for col, key in filter_cols.items():
            if col != exclude_col:
                sel = st.session_state[key]
                if isinstance(sel, str):
                    if sel != "الكل" and col in temp.columns:
                        temp = temp[temp[col] == sel]
                elif sel and col in temp.columns:  # list and not empty
                    temp = temp[temp[col].isin(sel)]
        return temp

    f1,f2,f3,f4,f5 = st.columns(5)

    # Build options for each filter from data filtered by others
    cat_filtered = get_filtered_excluding(df, filter_cols, "التصنيف")
    cat_options = sorted(cat_filtered["التصنيف"].dropna().unique()) if "التصنيف" in cat_filtered.columns else []
    cat_default = st.session_state.done_cat if isinstance(st.session_state.done_cat, list) else ([st.session_state.done_cat] if st.session_state.done_cat != "الكل" else [])
    cat = f1.multiselect("التصنيف", cat_options, default=cat_default, key="done_cat")

    mun_filtered = get_filtered_excluding(df, filter_cols, "البلدية")
    mun_options = sorted(mun_filtered["البلدية"].dropna().unique()) if "البلدية" in mun_filtered.columns else []
    mun_default = st.session_state.done_mun if isinstance(st.session_state.done_mun, list) else ([st.session_state.done_mun] if st.session_state.done_mun != "الكل" else [])
    mun = f2.multiselect("البلدية", mun_options, default=mun_default, key="done_mun")

    budget_filtered = get_filtered_excluding(df, filter_cols, "ميزانية المشروع")
    budget_options = sorted(budget_filtered["ميزانية المشروع"].dropna().unique()) if "ميزانية المشروع" in budget_filtered.columns else []
    budget_default = st.session_state.done_budget if isinstance(st.session_state.done_budget, list) else ([st.session_state.done_budget] if st.session_state.done_budget != "الكل" else [])
    budget = f3.multiselect("ميزانية المشروع", budget_options, default=budget_default, key="done_budget")

    year_filtered = get_filtered_excluding(df, filter_cols, "السنة")
    year_options = sorted(year_filtered["السنة"].dropna().unique()) if "السنة" in year_filtered.columns else []
    year_default = st.session_state.done_year if isinstance(st.session_state.done_year, list) else ([st.session_state.done_year] if st.session_state.done_year != "الكل" else [])
    year = f4.multiselect("السنة", year_options, default=year_default, key="done_year")

    project_filtered = get_filtered_excluding(df, filter_cols, "اسم المشروع")
    project_options = sorted(project_filtered["اسم المشروع"].dropna().unique()) if "اسم المشروع" in project_filtered.columns else []
    project_default = st.session_state.done_project if isinstance(st.session_state.done_project, list) else ([st.session_state.done_project] if st.session_state.done_project != "الكل" else [])
    project = f5.multiselect("اسم المشروع", project_options, default=project_default, key="done_project")

    # Apply all filters to get final filtered data
    filtered = get_filtered_excluding(df, filter_cols)

    # Calculate completed project counts from budget column
    num_bab3_completed = len(df[df["ميزانية المشروع"].astype(str).str.contains("الباب الثالث", na=False)]) if "ميزانية المشروع" in df.columns else 0
    num_bab4_completed = len(df[df["ميزانية المشروع"].astype(str).str.contains("الباب الرابع", na=False)]) if "ميزانية المشروع" in df.columns else 0

    # Calculate metrics
    total_contract = filtered["قيمة العقد"].sum() if "قيمة العقد" in filtered.columns else 0
    avg_progress = pd.to_numeric(filtered.get("نسبة الإنجاز", filtered.get("نسبة الانجاز", pd.Series())), errors="coerce").mean()

    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f"<div class='card blue'><h2>{total_contract:,.0f}</h2>قيمة العقود</div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='card green'><h2>{avg_progress:.1f}%</h2>نسبة الإنجاز</div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='card orange'><h2>{num_bab3_completed}</h2>مشاريع الباب الثالث المنجزة</div>", unsafe_allow_html=True)
    c4.markdown(f"<div class='card gray'><h2>{num_bab4_completed}</h2>مشاريع الباب الرابع المنجزة</div>", unsafe_allow_html=True)

    st.subheader("عدد المشاريع في كل بلدية")
    st.bar_chart(filtered["البلدية"].value_counts())

    st.subheader("تفاصيل المشاريع المنجزة")
    st.dataframe(filtered, use_container_width=True)

    # زر تحميل البيانات المفلترة كملف Excel باستخدام القالب
    excel_data = create_excel_from_template(filtered, TEMPLATE_PATH, LOGO_EXCEL_PATH, show_logo_in_excel, logo_excel_width)
    st.download_button(
        label="تحميل البيانات كExcel",
        data=excel_data,
        file_name=f"{st.session_state.top_nav.replace(' ', '_')}_filtered.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.stop()


# ================= الفلاتر (الباب الثالث + الرابع) =================
if st.session_state.top_nav in ["مشاريع الباب الثالث", "مشاريع الباب الرابع"]:

    st.subheader(f"تحليل {st.session_state.top_nav}")

    # Initialize filter states as lists
    if "bab_cat" not in st.session_state: st.session_state.bab_cat = []
    if "bab_ent" not in st.session_state: st.session_state.bab_ent = []
    if "bab_mun" not in st.session_state: st.session_state.bab_mun = []
    if "bab_stt" not in st.session_state: st.session_state.bab_stt = []
    if "bab_ct" not in st.session_state: st.session_state.bab_ct = []

    if st.button("إعادة تعيين الفلاتر"):
        st.session_state.bab_cat = []
        st.session_state.bab_ent = []
        st.session_state.bab_mun = []
        st.session_state.bab_stt = []
        st.session_state.bab_ct = []
        st.rerun()

    # Define filter columns based on section
    if st.session_state.top_nav == "مشاريع الباب الثالث":
        filter_cols = {
            "التصنيف": "bab_cat",
            "الجهة": "bab_ent",
            "البلدية": "bab_mun",
            "حالة المشروع": "bab_stt",
            "نوع العقد": "bab_ct"
        }
        num_cols = 5
    else:  # Bab4
        filter_cols = {
            "الجهة": "bab_ent",
            "البلدية": "bab_mun",
            "حالة المشروع": "bab_stt",
            "نوع العقد": "bab_ct"
        }
        num_cols = 4

    # Labels for filters
    labels = {
        "الجهة": "الجهة الرسمية"
    }

    def get_filtered_excluding_bab(df, filter_cols, exclude_col=None):
        temp = df.copy()
        for col, key in filter_cols.items():
            if col != exclude_col:
                sel = st.session_state[key]
                if isinstance(sel, str):
                    if sel != "الكل" and col in temp.columns:
                        temp = temp[temp[col] == sel]
                elif sel and col in temp.columns:
                    temp = temp[temp[col].isin(sel)]
        return temp

    f_cols = st.columns(num_cols)

    # Build multiselect for each filter
    for i, (col_name, key) in enumerate(filter_cols.items()):
        with f_cols[i]:
            filtered_ex = get_filtered_excluding_bab(df, filter_cols, col_name)
            options = sorted(filtered_ex[col_name].dropna().unique()) if col_name in filtered_ex.columns else []
            default_val = st.session_state[key] if isinstance(st.session_state[key], list) else ([st.session_state[key]] if st.session_state[key] != "الكل" else [])
            st.multiselect(labels.get(col_name, col_name), options, default=default_val, key=key)

    # Apply all filters
    filtered = get_filtered_excluding_bab(df, filter_cols)

# ================= KPI =================
k1,k2,k3,k4,k5,k6 = st.columns(6)

total_contract = filtered["قيمة العقد"].sum() if "قيمة العقد" in filtered.columns else 0
total_claims = filtered["قيمة المستخلصات"].sum() if "قيمة المستخلصات" in filtered.columns else 0
total_remain = filtered["المتبقي من المستخلص"].sum() if "المتبقي من المستخلص" in filtered.columns else 0
spend_ratio = (total_claims / total_contract * 100) if total_contract > 0 else 0

progress_ratio = 0
if "نسبة الإنجاز" in filtered.columns and "قيمة العقد" in filtered.columns:
    w = filtered.dropna(subset=["قيمة العقد","نسبة الإنجاز"])
    if not w.empty:
        progress_ratio = (w["قيمة العقد"] * w["نسبة الإنجاز"]).sum() / w["قيمة العقد"].sum()

k1.markdown(f"<div class='card blue'><h2>{len(filtered)}</h2>عدد المشاريع</div>", unsafe_allow_html=True)
k2.markdown(f"<div class='card green'><h2>{total_contract:,.0f}</h2>قيمة العقود</div>", unsafe_allow_html=True)
k3.markdown(f"<div class='card gray'><h2>{total_claims:,.0f}</h2>المستخلصات</div>", unsafe_allow_html=True)
k4.markdown(f"<div class='card orange'><h2>{total_remain:,.0f}</h2>المتبقي</div>", unsafe_allow_html=True)
k5.markdown(f"<div class='card blue'><h2>{spend_ratio:.1f}%</h2>نسبة الصرف</div>", unsafe_allow_html=True)
k6.markdown(f"<div class='card green'><h2>{progress_ratio:.1f}%</h2>نسبة الإنجاز</div>", unsafe_allow_html=True)

# ================= حالة المشاريع =================
st.subheader("حالة المشاريع")
st.markdown('<div class="chart-grid">', unsafe_allow_html=True)

# Chart 1
st.markdown('<div class="chart-item">', unsafe_allow_html=True)
sdf = build_status_df(filtered)
st.altair_chart(
    alt.Chart(sdf).mark_bar().encode(
        x="عدد",
        y=alt.Y("الحالة", sort="-x"),
        color=alt.Color("الحالة", scale=alt.Scale(domain=sdf["الحالة"], range=sdf["لون"]))
    ),
    use_container_width=True
)
st.markdown('</div>', unsafe_allow_html=True)

# Chart 2
st.markdown('<div class="chart-item">', unsafe_allow_html=True)
st.subheader("عدد المشاريع حسب البلدية")
st.bar_chart(filtered["البلدية"].value_counts())
st.markdown('</div>', unsafe_allow_html=True)

# Chart 3
st.markdown('<div class="chart-item">', unsafe_allow_html=True)
st.subheader("عدد المشاريع حسب حالة المشروع")
st.bar_chart(filtered["حالة المشروع"].value_counts())
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

# ================= التنبيهات =================
st.subheader("تنبيهات المشاريع")

overdue = filtered[filtered["حالة المشروع"].astype(str).str.contains("متأخر|متعثر", na=False)]

if "تاريخ الانتهاء" in filtered.columns and "نسبة الإنجاز" in filtered.columns:
    risk = filtered[
        (filtered["تاريخ الانتهاء"] <= pd.Timestamp.today() + timedelta(days=30)) &
        (filtered["نسبة الإنجاز"] < 70)
    ]
else:
    risk = pd.DataFrame()

# إضافة سبب التوقع للتأخير في جدول المشاريع المتوقع تأخرها
if not risk.empty:
    risk = risk.copy()
    risk["سبب التوقع للتأخير"] = "التاريخ المتبقي أقل من 30 يوماً والإنجاز أقل من 70%"

b1,b2 = st.columns(2)
if b1.button(f"المشاريع المتأخرة ({len(overdue)})"):
    st.session_state.show_overdue = not st.session_state.show_overdue
if b2.button(f"المشاريع المتوقع تأخرها ({len(risk)})"):
    st.session_state.show_risk = not st.session_state.show_risk

if st.session_state.show_overdue:
    st.dataframe(overdue, use_container_width=True)
    excel_data_overdue = create_excel_from_template(overdue, TEMPLATE_PATH, LOGO_EXCEL_PATH, show_logo_in_excel, logo_excel_width)
    st.download_button(
        label="تحميل المشاريع المتأخرة كExcel",
        data=excel_data_overdue,
        file_name="overdue_projects.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
if st.session_state.show_risk:
    st.dataframe(risk, use_container_width=True)
    excel_data_risk = create_excel_from_template(risk, TEMPLATE_PATH, LOGO_EXCEL_PATH, show_logo_in_excel, logo_excel_width)
    st.download_button(
        label="تحميل المشاريع المتوقع تأخرها كExcel",
        data=excel_data_risk,
        file_name="risk_projects.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ================= جدول =================
st.markdown("---")
st.subheader("تفاصيل المشاريع")

# زر تحميل البيانات المفلترة كملف Excel باستخدام القالب
excel_data = create_excel_from_template(filtered, TEMPLATE_PATH, LOGO_EXCEL_PATH, show_logo_in_excel, logo_excel_width)
st.download_button(
    label="تحميل البيانات كExcel",
    data=excel_data,
    file_name=f"{st.session_state.top_nav.replace(' ', '_')}_filtered.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.dataframe(filtered, use_container_width=True)
